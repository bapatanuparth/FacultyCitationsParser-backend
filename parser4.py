import textract
import re
import spacy
import openpyxl
from parser1 import process_citation

import fitz


# Load spaCy's English-language model, assuming it's already installed
from docx import Document
import spacy

# Load spaCy model
nlp = spacy.load("en_core_web_sm")

ambiguous=["PROFESSIONAL DEVELOPMENT WORKSHOPS","Publication", "Editorial Activities" ]

allowed_headers=["Book chapters:","Book Chapters","Book Reviews","Published Conference Proceedings","Other Publications"]

publication_headers = [
    "Publications", "Journal Articles", "Published Articles", "Published Works","Published Papers",
   "Published Research", "Research Papers", "Scientific Publications", 
    "Review Articles",
   "Peer reviewed journals ","Refereed Journals"
]

book_headers= ["Books and Book Chapters", "Books", "Book chapters","Book reviews", "Book chapter"]

conference_headers=["Conference Proceedings","Conference presentations", "Conference Workshops","Presentations", "Conference papers","PEER REVIEWED PROCEEDINGS"]

exception_dict = {
    "book chapters and other publications": "Book Chapter",
    "book chapters & other publications":"Book Chapter"
    # Add more exceptions here as needed
}


def classify_header_type(header):
    # Combine all keywords into a dictionary with categories
    header_types = {
        "Publication": publication_headers,
        "Book Chapter": book_headers,
        "Conference": conference_headers,
    }
    
    # Lowercase the header for case-insensitive comparison
    header_lower = header.lower()
    
    if header_lower in exception_dict:
        return exception_dict[header_lower]
    # Check each category for a match
    for category, keywords in header_types.items():
        if any(keyword.lower() in header_lower for keyword in keywords):
            return category
    
    # Return None or a default category if no match is found
    return None

def remove_text_in_parentheses(line):
    """
    Removes text within parentheses, including the parentheses themselves.
    """
    # Regular expression to match content within parentheses
    new_line = re.sub(r'\(.*?\)', '', line)
    return new_line.strip()

def is_header(line):
    cleaned_line=remove_text_in_parentheses(line)
    doc = nlp(cleaned_line)
    all_first_letters_capitalized = all(word[0].isupper() for word in cleaned_line.split() if word)
    is_short_cleaned_line = len(cleaned_line.split()) <= 6  # Adjust based on your criteria for short cleaned_line
    all_tokens_are_proper_nouns_or_nouns = all(token.pos_ in ['PROPN', 'NOUN'] for token in doc)
    contains_no_numbers = all(not any(char.isdigit() for char in word) for word in cleaned_line.split())
    contains_no_asterisk = '*' not in cleaned_line 

    if (all_first_letters_capitalized or is_short_cleaned_line or all_tokens_are_proper_nouns_or_nouns) and contains_no_numbers and contains_no_asterisk:
        return True
    return False

def filter_records_by_year(publications, conferences, book_chapters, year="2023"):
    year_pattern = re.compile(rf'\b{year}\b')

    def filter_by_year(entries):
        return [entry for entry in entries if year_pattern.search(entry)]

    filtered_publications = filter_by_year(publications)
    filtered_conferences = filter_by_year(conferences)
    filtered_book_chapters = filter_by_year(book_chapters)

    return filtered_publications, filtered_conferences, filtered_book_chapters

def write_sections_to_file(publications, conferences, book_chapters, output_file_path,header_types):
    with open(output_file_path, 'w', encoding='utf-8') as file:
        for header, type in header_types:
            file.write(f"Header: {header}, Type: {type}\n")
        # Write Publications section
        file.write("Publications:\n")
        for publication in publications:
            file.write(f"- {publication}\n")
        file.write("\n")
        
        # Write Conferences section
        file.write("Conferences:\n")
        for conference in conferences:
            file.write(f"- {conference}\n")
        file.write("\n")
        
        # Write Book Chapters section
        file.write("Book Chapters:\n")
        for book_chapter in book_chapters:
            file.write(f"- {book_chapter}\n")
        file.write("\n")

def find_person_name(resume_text):
    # Process the resume text with spaCy
    lines = resume_text.split("\n")
    excluded_words = {'curriculum', 'vita', 'resume', 'cv'}
    # Keep only the first two non-empty lines
    # first_two_lines = [line for line in lines if line.strip()][:3]
    first_three_lines = []
    for line in lines:
        if line.strip() and not any(word in line.lower() for word in excluded_words):
            first_three_lines.append(line)
            if len(first_three_lines) == 3:
                break
    
    # Process these lines with spaCy and collect named entities
    named_entities = []
    for line in first_three_lines:
        print("printint lines" + line)
        doc = nlp(line)
        for ent in doc.ents:
            print(ent.text + " " + ent.label_)
            named_entities.append((ent.text, ent.label_))
            if ent.label_ == "PERSON":
                return ent.text  # Return the first 'PERSON' entity found
    
    # If no 'PERSON' entity is found, attempt to return text not recognized as any named entity
    for line in first_three_lines:
        doc = nlp(line)  # Re-process each line to check for entities
        if not doc.ents:  # If the line has no named entities, consider it as the person's name
            return line.strip()

    # If no name is found by now, return a default message
    return "Name not found"

def process_resume(file_path):
    if is_pdf(file_path):
       
        doc = fitz.open(file_path)
        text = '\n'.join([page.get_text() for page in doc])
    else:
       
        text = textract.process(file_path).decode('utf-8')
    name= find_person_name(text)
    # Split the text into lines
    lines = text.split('\n')
    current_section_type = None
    sections = {
        "Publication": [],
        "Conference": [],
        "Book Chapter": [],
    }
    header_types = []
    # print(lines)

    for line in lines:
        line = line.strip()
        if line and is_header(line):
            # Determine the type of the header
            header_type = classify_header_type(line)
            print(line)
            if header_type in sections:
                current_section_type = header_type
                header_types.append((line, header_type)) 
            else:
                # If the header is not one of the specified types, reset the current section type
                current_section_type = None
        elif current_section_type:
            # If we are within a recognized section, append the line to the appropriate list
            
            sections[current_section_type].append(line)

    return sections["Publication"], sections["Conference"], sections["Book Chapter"], header_types, name

def remove_text_in_parentheses_and_digits(line):
    """
    Removes text within parentheses and all digits from the line.
    """
    # Remove content within parentheses
    line_no_parentheses = re.sub(r'\(.*?\)', '', line)
    # Remove digits
    cleaned_line = re.sub(r'\d+', '', line_no_parentheses)
    return cleaned_line.strip()

def is_header_capital(line):
    """
    Checks if a cleaned line (without text in parentheses and digits) is written in all capital letters,
    or matches an entry in allowed_headers, even if not in all caps.
    """
    cleaned_line = remove_text_in_parentheses_and_digits(line)
    # Check if the cleaned line is in all caps or matches an allowed header
    # Also ensure the line is not empty after cleaning
    return cleaned_line.isupper() and cleaned_line != "" or cleaned_line in allowed_headers

def process_resume_headers_capital(file_path):
    if is_pdf(file_path):
        doc = fitz.open(file_path)
        text = '\n'.join([page.get_text() for page in doc])
    else:
        text = textract.process(file_path).decode('utf-8')
    name= find_person_name(text)
    # Split the text into lines
    lines = text.split('\n')
    current_section_type = None
    sections = {
        "Publication": [],
        "Conference": [],
        "Book Chapter": [],
    }
    header_types = []
    # print(lines)

    for line in lines:
        line = line.strip()
        if line and is_header_capital(line):
            # Determine the type of the header
            header_type = classify_header_type(line)
            print(line)
            if header_type in sections:
                current_section_type = header_type
                header_types.append((line, header_type)) 
            else:
                # If the header is not one of the specified types, reset the current section type
                current_section_type = None
        elif current_section_type:
            # If we are within a recognized section, append the line to the appropriate list
            
            sections[current_section_type].append(line)

    return sections["Publication"], sections["Conference"], sections["Book Chapter"], header_types, name

def populate_sheet(ws, data, name, extracted_data):
    # Assuming row 1 is headers: Name, Publication/Conference/Book Chapter, Authors, Journal Name/Presentation Venue, Paper Title, Year
    row = ws.max_row + 1  # Start writing from the next empty row
    for item, details in zip(data, extracted_data):
        is_elite = 'Yes' if details.get('elite_journal') else 'No'
        ws[f'A{row}'] = name
        ws[f'B{row}'] = item  # The whole citation text
        ws[f'C{row}'] = ', '.join(details['authors'])
        ws[f'D{row}'] = details['paper_title']
        ws[f'E{row}'] = details['journal_name']
        ws[f'F{row}'] = details['year']
        ws[f'G{row}'] = is_elite
        row += 1

def populate_excel(output_path, filtered_publications, filtered_conferences, filtered_book_chapters, name, extracted_pubs, extracted_conf,extracted_books):
    wb = openpyxl.load_workbook(output_path)
    
    # Populate the 'Publications' sheet
    if "Publications" in wb.sheetnames:
        ws = wb["Publications"]
        populate_sheet(ws, filtered_publications, name, extracted_pubs)
    
    # Populate the 'Conference' sheet
    if "Conference" in wb.sheetnames:
        ws = wb["Conference"]
        populate_sheet(ws, filtered_conferences, name, extracted_conf)
    
    # Populate the 'Book Chapters' sheet
    if "Book Chapters" in wb.sheetnames:
        ws = wb["Book Chapters"]
        populate_sheet(ws, filtered_book_chapters, name, extracted_books)  # Assuming no extracted data for book chapters yet

    if "Elite Journal Submissions" in wb.sheetnames:
        ws=wb["Elite Journal Submissions"]
        populate_elite_journals(ws, name,extracted_pubs)
        populate_elite_journals(ws, name,extracted_conf)
        populate_elite_journals(ws, name,extracted_books)
    
    wb.save(output_path)

def populate_elite_journals(ws_elite, name, extracted_data):
    row = ws_elite.max_row + 1 
    for pub in extracted_data:
        elite_=pub.get('elite_journal')
        if elite_:  # Check if the publication is from an elite journal
            ws_elite[f'A{row}'] = name
            ws_elite[f'B{row}'] = pub['citation']  # The whole citation text
            ws_elite[f'C{row}'] = pub['elite_journal']  # The name of the elite journal
            row += 1  # Increment row for the next entry

# def populate_excel(output_path,filtered_publications,filtered_conferences,filtered_book_chapters,name,extracted_pubs,extracted_conf):
#     # Load or create an Excel workbook
#     wb = openpyxl.load_workbook(output_path)
    
#     # Populate the 'Publications' sheet
#     if "Publications" in wb.sheetnames:
#         ws = wb["Publications"]
#         populate_sheet(ws, "Publications", filtered_publications,name)
    
#     # Populate the 'Conference' sheet
#     if "Conference" in wb.sheetnames:
#         ws = wb["Conference"]
#         populate_sheet(ws, "Conference Presentations", filtered_conferences,name)
    
#     # Populate the 'Book Chapters' sheet
#     if "Book Chapters" in wb.sheetnames:
#         ws = wb["Book Chapters"]
#         populate_sheet(ws, "Book Chapters", filtered_book_chapters,name)
    
#     # Save the workbook
#     wb.save(output_path)

def is_pdf(file_path):
    return file_path.lower().endswith('.pdf')

# def populate_sheet(ws, column_header, items,name):
#     # Find column indices based on headers
#     name_col = ws['A']
#     content_col = ws['B']
    
#     # Start writing from the first empty row
#     row_number = len(name_col) + 1
    
#     for item in items:
#         ws[f'A{row_number}'] = name
#         ws[f'B{row_number}'] = item
#         row_number += 1

def start(file_path,headers_capital,year):
    # Example usage
    # Update with your file path
    if headers_capital:
        publications, conferences, book_chapters, header_types,name  = process_resume_headers_capital(file_path)
    else:
        publications, conferences, book_chapters, header_types,name  = process_resume(file_path)
    # print(publications)
    # Example call to filter records for the year 2023
    filtered_publications, filtered_conferences, filtered_book_chapters = filter_records_by_year(publications, conferences, book_chapters, year)
    print(name)
    print(filtered_publications)
    print(filtered_conferences)
    print(filtered_book_chapters)

    extracted_pubs=process_citation(filtered_publications)
    extracted_conf=process_citation(filtered_conferences)
    extracted_books=process_citation(filtered_book_chapters)

        # Call the function with the path to your output file
    output_file_path = './sections.txt'  # Update with your actual output file path
    write_sections_to_file(publications, conferences, book_chapters, output_file_path,header_types)

    # output_excel_path = 'C:\WORK\PatentsView_data\ResumeParser\Faculty research publications details TEST.xlsx'
    # populate_excel(output_excel_path,filtered_publications, filtered_conferences, filtered_book_chapters,name,extracted_pubs,extracted_conf,extracted_books)
    combined_data = {
        "name":name,
        "publications": extracted_pubs,
        "conferences": extracted_conf,
        "bookChapters":extracted_books
    }
    return combined_data


file_path = r"C:\Users\Parth\Downloads\General\General\Tenure Track\Townsend, David\Townsend CV Winter 2024.pdf"
# start(file_path,True)